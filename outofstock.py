import openpyxl
import requests
import os
import time
import random
import logging
import time
import datetime
import win32com.client as win32
import re
from tqdm import tqdm
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

def clean_asin(asin):
    # Remove leading and trailing spaces
    asin = asin.strip()
    
    # Remove special characters
    asin = re.sub(r'[^A-Za-z0-9]+', '', asin)
    
    return asin

def format_time(seconds):
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{int(hours)}h {int(minutes)}m {int(seconds)}s"

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36 Edge/17.17134',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/11.0 Safari/605.1.15',
    'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:15.0) Gecko/20100101 Firefox/15.0.1',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:60.0) Gecko/20100101 Firefox/60.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:61.0) Gecko/20100101 Firefox/61.0',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:59.0) Gecko/20100101 Firefox/59.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/12.1.2 Safari/605.1.15',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:54.0) Gecko/20100101 Firefox/54.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_2) AppleWebKit/601.3.9 (KHTML, like Gecko) Version/9.0.2 Safari/601.3.9',
    'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.3',
    'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.84 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:57.0) Gecko/20100101 Firefox/57.0',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.3',
    'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:52.0) Gecko/20100101 Firefox/52.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2227.1 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.75 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.102 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; AS; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64; rv:68.0) Gecko/20100101 Firefox/68.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.13; rv:63.0) Gecko/20100101 Firefox/63.0',
    'Mozilla/5.0 (Windows NT 6.3; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.102 Safari/537.36 Edge/18.19582',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.84 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.167 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36',
    'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:71.0) Gecko/20100101 Firefox/71.0',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36',
    'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36',
    
]

def read_asins_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    asins = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        asin = row[0]
        cleaned_asin = clean_asin(asin)
        asins.append(cleaned_asin)
        asins.append(asin)

    return asins


# Check the availability of a product by URL
def check_availability(url, region):
    headers = {
        'User-Agent': random.choice(USER_AGENTS)
    }
    for _ in range(3):  # Retry up to 3 times
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'lxml')

            out_of_stock = soup.find('span', {'class': 'a-size-medium a-color-price'}) or \
                           soup.find('div', {'id': 'availability'}) or \
                           soup.find('span', {'class': 'a-size-medium a-color-state'})

            out_of_stock_phrases = get_out_of_stock_phrases(region)
            if out_of_stock and any(phrase in out_of_stock.text.lower() for phrase in out_of_stock_phrases):
                return 'Out Of Stock'
            else:
                return 'In Stock'

        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 404:
                logging.exception(f"URL not found: {url}")
                return 'URL not found'
            else:
                logging.exception(f"HTTPError for URL: {url}")
                time.sleep(5)
    return 'Failed to fetch URL'


def generate_urls_for_asin(asin):
    domains = {
        'UK': 'https://www.amazon.co.uk/dp/',
        'ES': 'https://www.amazon.es/dp/',
        'FR': 'https://www.amazon.fr/dp/',
        'DE': 'https://www.amazon.de/dp/',
        'IT': 'https://www.amazon.it/dp/'
    }
    # Check for None values and print them
    for region, domain in domains.items():
        if domain is None:
            print(f"Domain for region {region} is None")

    urls = {region: domain + asin for region, domain in domains.items() if domain is not None and asin is not None}
    return urls

def get_out_of_stock_phrases(region):
    phrases = {
        'UK': ['out of stock', 'temporarily out of stock'],
        'ES': ['agotado temporalmente', 'out of stock'],
        'FR': ['temporairement en rupture de stock', 'out of stock'],
        'DE': ['derzeit nicht auf lager', 'temporarily out of stock', 'out of stock'],
        'IT': ['al momento non disponibile', 'out of stock'],
    }
    return phrases.get(region, [])

# Determine the region based on the URL
def get_region_from_url(url):
    if 'amazon.es' in url:
        return 'ES'
    elif 'amazon.fr' in url:
        return 'FR'
    elif 'amazon.de' in url:
        return 'DE'
    elif 'amazon.it' in url:
        return 'IT'
    elif 'amazon.co.uk' in url:
        return 'UK'

# Write the results to a new output Excel file
# Write the results to a new output Excel file
def write_results_to_excel(output_file_path, results):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the header row
    sheet.cell(row=1, column=1).value = 'ASIN'
    sheet.cell(row=1, column=2).value = 'UK'
    sheet.cell(row=1, column=3).value = 'DE'
    sheet.cell(row=1, column=4).value = 'FR'
    sheet.cell(row=1, column=5).value = 'ES'
    sheet.cell(row=1, column=6).value = 'IT'

    # Write the results
    for index, (asin, asin_result) in enumerate(results.items(), start=2):
        sheet.cell(row=index, column=1).value = asin
        for col, region in enumerate(('UK', 'DE', 'FR', 'ES', 'IT'), start=2):
            sheet.cell(row=index, column=col).value = asin_result[region]

    workbook.save(output_file_path)


def status_text(status):
    if status == "Out Of Stock":
        return "OOS"
    elif status == "URL not found":
        return "No"
    else:
        return "Yes"


def load_or_create_workbook(file_path):
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.save(file_path)
        workbook = load_workbook(file_path)
    return workbook

def send_email(subject, body, attachment_path):
    outlook = win32.Dispatch('outlook.application')
    namespace = outlook.GetNamespace("MAPI")

    mail = outlook.CreateItem(0)
    mail.Subject = subject
    mail.Body = body
    mail.Attachments.Add(attachment_path)

    # Add recipients (To and CC)
    mail.To = "ritchie.emery@funko.com"  # Replace with the recipient's email address
    mail.CC = "gabriel.konopnicki@funko.com"  # Replace with the CC recipient's email address

    # Use the following line to display the email before sending
    mail.Display()

    # Uncomment the following line to send the email
    # mail.Send()

    print("Email created in Outlook")

def get_formatted_date():
    today = datetime.now()
    return today.strftime("ASINs Status %d%b%Y")

def create_email_body(results, total_urls, total_asins):
    out_of_stock_count = sum(1 for asin_results in results.values() for region_status in asin_results.values() if region_status == 'No (OOS)')
    suppressed_count = sum(1 for asin_results in results.values() for region_status in asin_results.values() if region_status == 'No (Suppressed)')

    body = f"""\
    <html>
    <head></head>
    <body>
        <h1>ASINs Status Report</h1>
        <p>Total ASINs: {total_asins}</p>
        <p>Total URLs processed: {total_urls}</p>
        <p>Out of Stock ASINs: {out_of_stock_count}</p>
        <p>Suppressed ASINs: {suppressed_count}</p>
        <p>Please find the attached Excel file for the full results.</p>
    </body>
    </html>
    """
    return body

    
# Main function
def main():
    total_asins_processed = 0
    input_file_path = r'C:\Users\gabriel.konopnicki\OneDrive - funko.com\Desktop\input\list.xlsx'

    # Generate the output file name based on today's date
    today = datetime.today().strftime('%d%b%Y')
    output_file_name = f"ASINS_Status_{today}.xlsx"
    output_file_path = fr'C:\Users\gabriel.konopnicki\OneDrive - funko.com\Desktop\input\output\{output_file_name}'

    asins = read_asins_from_excel(input_file_path)

    results = {asin: {'UK': '', 'DE': '', 'FR': '', 'ES': '', 'IT': ''} for asin in asins}
    total_urls = len(asins) * 5  # Assuming 5 domains for each ASIN
    total_asins = len(asins) 
    urls_processed = 0
    start_time = time.time()

    # Initialize the progress bar
    progress_bar = tqdm(total=total_urls, desc="Processing", unit="URL", ncols=125)

    # Check the availability for each ASIN and each Amazon domain
    for asin in asins:
        urls = generate_urls_for_asin(asin)
        for region, url in urls.items():
            status = check_availability(url, region)
            results[asin][region] = status_text(status)

            # Calculate progress information
            urls_processed += 1
            progress = (urls_processed / total_urls) * 100
            elapsed_time = time.time() - start_time
            time_per_url = elapsed_time / urls_processed
            time_remaining = (total_urls - urls_processed) * time_per_url

            # Update the progress bar
            progress_bar.set_postfix({"ASIN": asin, "Region": region, "Status": status, "Time remaining": format_time(time_remaining)}, refresh=False)
            progress_bar.update(1)

            # Log progress information
            logging.info(f"ASIN: {asin}, Region: {region}, URL: {url}, Status: {status}, Progress: {progress:.2f}%, Time remaining: {format_time(time_remaining)}")

            time.sleep(10)  # Add a delay of 5 seconds between requests
        total_asins_processed += 1

    # Load or create the output workbook and write the results
    load_or_create_workbook(output_file_path)
    write_results_to_excel(output_file_path, results)

    # Create and send an email with the output file attached
    subject = get_formatted_date()
    body = create_email_body(results, total_urls, total_asins)
    send_email(subject, body, output_file_path)

    # Print the success message
    print("Hooray! The process has been successfully completed.")

# Run the main function
if __name__ == '__main__':
    main()
